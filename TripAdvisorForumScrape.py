import openpyxl, os
import requests
from lxml import html
import time,re

'''Fetches all url from the excel file from "SheetName"'''

# workbook = openpyxl.load_workbook(excelPath)
# worksheet = workbook.get_sheet_by_name(SheetName)
filename='NYC _100_Link_To_FreeLancer.xlsx'
excelPath=(os.path.dirname(__file__))+'/'+filename
print(excelPath)
workbook = openpyxl.load_workbook(excelPath)
worksheet = workbook.get_active_sheet()

website = []
for row in worksheet.iter_rows(min_row=1, min_col=3, max_col=3):  ##Getting all the websites,102 max row
    for cell in row:
        website.append(cell.value)

'''End of first block'''


class TripAdvisorXpathForUser1:
    """First Website"""

    def __init__(self, seq_index00):
        self.baseUrl='https://www.tripadvisor.in'
        self.seqIndex = seq_index00
        self.pageContent = requests.get(website[self.seqIndex - 1])
        self.tree = html.fromstring(self.pageContent.content)
        self.Title = self.tree.xpath('//div[@class="firstPostBox"]//div[@class="postTitle"]/span/text()')
        self.User1 = self.tree.xpath('//div[@class="username"]/a/span/text()')
        self.xPath1 = '//span[@class="postNum" and contains(text(),\"'
        self.user1xpath2 = '.\")]/../../../..//div[@class="username"]/a/span/text()'
        self.date1xPath2 = '.\")]/../../../..//div[@class="postDate"]/text()'  # v02
        self.location1xPath2 = '.\")]/../../../..//div[@class="username"]/..//div[@class="location"]/text()'
        self.posts1xPath2 = '.\")]/../../../..//div[@class="postBadge badge"]/span/text()'
        self.reviewsxPath2 = '.\")]/../../../..//div[@class="reviewerBadge badge"]/span/text()'
        self.replies1xPath2 = '.\")]/../../../..//div[@class="postBody"]/text()'
        self.User1 = self.tree.xpath('//div[@class="username"]/a/span/text()')
        self.Date1 = self.tree.xpath('//div[@class="postDate"]/text()')
        self.Location1 = self.tree.xpath('//div[@class="username"]/..//div[@class="location"]/text()')
        self.Posts1 = self.tree.xpath('//div[@class="postBadge badge"]/span/text()')
        self.Reviews1 = self.tree.xpath('//div[@class="reviewerBadge badge"]/span/text()')
        self.Replies1 = self.tree.xpath('//div[@class="postBody"]')
        self.Rep1 = []
        for i in range(len(self.Replies1)):
            self.Rep1.append(re.sub(r'[\n\t]','',self.Replies1[i].text_content()))
        self.firstPostValues = {}
        self.forumUsers={}
        self.forumDates={}
        self.forumLocation={}
        self.forumPosts={}
        self.forumReviews={}
        self.forumReplies={}
        #NextPageVariables
        self.Rep2 = []

    def firstPost(self):
        #firstPostValues = {}
        print("Printing the FirstUser Post")
        self.firstPostValues = {1: self.Title[0],
                           2: self.User1[0],
                           3: self.Date1[0],
                           4: self.Location1[0],
                           5: re.sub(r'[a-zA-Z]|[A-Z]* ','',self.Posts1[0],),
                           6: re.sub(r'[a-zA-Z]|[A-Z]* ','',self.Reviews1[0]),
                           7: self.Rep1[0]
                           }
        print("Printing the FirstUser Post")
        for index in range(1,8):
            colnumber = 3+index
            try:
                worksheet.cell(row=self.seqIndex, column=colnumber).value = self.firstPostValues[index]
            except KeyError as e:
                    print("Unable to print value becuase of %s and at index %d"%(e,index))
        print(self.seqIndex,self.firstPostValues, sep="|", end="\n")

#######
    def forumReplyPost(self):
        print("Printing Community Post")
        print("Forum Users")
        #forumUsers = {}
        for seq in range(1, len(self.Date1)):  # len(Date1))
            # print("seq", seq)
            try:
                if not self.tree.xpath(self.xPath1 + str(seq) + self.user1xpath2):  # == '':
                    # print("if", seq)
                    # print("Invalid xpath with no value")
                    self.forumUsers[seq] = "No"
                else:
                    # print("else", seq)
                    self.forumUsers[seq] = self.tree.xpath(self.xPath1 + str(seq) + self.user1xpath2)
            except IndexError as e:
                print("IndexError occurred at Last Index seq in Users was", seq)
                self.forumUsers[seq] = "No"
                continue
            except:
                print("Except Seq", seq)
        #print("In the End, Users are", Users)

        print("Forum Dates")
        #forumDates = {}
        for seq in range(1, len(self.Date1)):  # len(Date1))
            # print("seq", seq)
            try:
                if not self.tree.xpath(self.xPath1 + str(seq) + self.date1xPath2):
                    # print("if", seq)
                    # print("Invalid xpath with no value")
                    #print(self.tree.xpath(self.xPath1 + str(seq) + self.date1xPath2))
                    self.forumDates[seq] = "No"
                else:
                    # print("else", seq)
                    self.forumDates[seq] = self.tree.xpath(self.xPath1 + str(seq) + self.date1xPath2)
            except IndexError as e:
                print("IndexError occurred at Last Index seq in Dates was", seq)
                self.forumDates[seq] = "No"
                continue
            except:
                print("Dates:Except Seq", seq)
        #print("In the End, Dates are", Dates)

        print("Forum Locations")
        #forumLocation = {}
        for seq in range(1, len(self.Date1)):
            try:
                if not self.tree.xpath(self.xPath1 + str(seq) + self.location1xPath2):  # == '':
                    self.forumLocation[seq] = "No"
                else:
                    self.forumLocation[seq] = self.tree.xpath(self.xPath1 + str(seq) + self.location1xPath2)
            except IndexError as e:
                print("IndexError occurred at Last Index seq in location1 was", seq)
                self.forumLocation[seq] = "No"
                continue
            except:
                print("UnExepected in location", seq)

        print("Forum posts1")
        #forumPosts = {}
        for seq in range(1, len(self.Date1)):
            try:
                if not self.tree.xpath(self.xPath1 + str(seq) + self.posts1xPath2):  # == '':
                    # print("if", seq)
                    # print("Invalid xpath with no value")
                    self.forumPosts[seq] = "No"
                else:
                    # print("else", seq)
                    self.forumPosts[seq] = re.sub(r'[a-zA-Z]|[A-Z]* ','',str(self.tree.xpath(self.xPath1 + str(seq) + self.posts1xPath2)))
            except IndexError as e:
                print("Last Index seq was", seq)
                self.forumPosts[seq] = "No"
                continue
            # except:
            #     print("Except Seq", seq)

        print("Forum reviews")
        #forumReviews = {}
        for seq in range(1, len(self.Date1)):  # len(Date1))
            # print("seq", seq)
            try:
                if not self.tree.xpath(self.xPath1 + str(seq) + self.reviewsxPath2):  # == '':
                    # print("if", seq)
                    # print("Invalid xpath with no value")
                    self.forumReviews[seq] = "No"
                else:
                    # print("else", seq)
                    self.forumReviews[seq] = re.sub(r'[a-zA-Z]|[A-Z]* ','',str(self.tree.xpath(self.xPath1 + str(seq) + self.reviewsxPath2)))
            except IndexError as e:
                print("Last Index seq was", seq)
                self.forumReviews[seq]="No"
                continue
            except:
                print("Except Seq", seq)
        print("Forum replies1")
        #forumReplies = {}
        for seq in range(1, len(self.Date1)):
            try:
                if not self.Rep1[seq]:
                    print("if", seq)
                    print("Invalid xpath with no value")
                    self.forumReplies[seq] = "No"
                else:
                    # print("else", seq)
                    self.forumReplies[seq] = self.Rep1[seq]
            except IndexError as e:
                print("IndexError occurred at Last Index seq in replies1 was", seq)
                self.forumReplies[seq] = "No"
                continue
            except:
                print("Except Seq", seq)
        print("Insertion in excel started")
        for index, colnumber in zip((range(1, len(self.Date1))), range(11, 11 + (len(self.Date1) * 6), 6)):
            try:
                print("formuReplyPost", index, colnumber, self.forumUsers[index], self.forumDates[index],
                      self.forumLocation[index],self.forumPosts[index], self.forumReviews[index],
                      self.forumReplies[index], sep="|", end="\n")
                worksheet.cell(row=self.seqIndex, column=colnumber + 0).value = (re.sub(r"[\[\'*\'\]]",'',   str(self.forumUsers[index])))
                worksheet.cell(row=self.seqIndex, column=colnumber + 1).value = (re.sub(r"[\[\'*\'\]]",'',   str(self.forumDates[index])))
                worksheet.cell(row=self.seqIndex, column=colnumber + 2).value = (re.sub(r"[\[\'*\'\]]",'',   str(self.forumLocation[index])))
                worksheet.cell(row=self.seqIndex, column=colnumber + 3).value = (re.sub(r"[\[\'*,*\'\]]",'',   str(self.forumPosts[index])))
                worksheet.cell(row=self.seqIndex, column=colnumber + 4).value = (re.sub(r"[\[\'*\'\]]",'',   str(self.forumReviews[index])))
                worksheet.cell(row=self.seqIndex, column=colnumber + 5).value = (re.sub(r'[\[\]]','',   str(self.forumReplies[index])))
                # print("formuReplyPost", index, colnumber, self.forumUsers[index], self.forumDates[index],
                #       self.forumLocation[index], self.forumPosts[index], self.forumReviews[index],
                #       self.forumReplies[index], sep="|", end="\n")
            except IndexError:
                print("IndexErrorException unhandled at index %d while insertion"%(index))
                continue
            except KeyError as KE:
                print("Dictionay cannot reference this key at index:%d",index,KE)
    def NextPageforumReplyPost(self):
        if not self.tree.xpath('.//*[@id="pager_top2"]/a[2]/@href'):
            print("Next Page not Present")
        else:
            NextPage = self.tree.xpath('.//*[@id="pager_top2"]/a[2]/@href')
            print("NextPage", NextPage[0])
            pageContent2 = requests.get(self.baseUrl + NextPage[0])
            tree2 = html.fromstring(pageContent2.content)
            startIndexNextPage=11
            User2 = tree2.xpath('//div[@class="username"]/a/span/text()')
            Date2 = tree2.xpath('//div[@class="postDate"]/text()')
            Replies2 = tree2.xpath('//div[@class="postBody"]')
            for i in range(1,len(Replies2)):
                self.Rep2.append(re.sub(r'[\n\t]', '', Replies2[i].text_content()))
            print("User2", User2,Date2)
            print("Printing Community Post of next page")
            print("Forum Users of next page")
            # forumUsers = {}
            for seq in range(startIndexNextPage, len(Date2)+startIndexNextPage-1-1):  # len(Date1))
                print("seq", seq)
                try:
                    if not tree2.xpath(self.xPath1 + str(seq) + self.user1xpath2):  # == '':
                        # print("if", seq)
                        # print("Invalid xpath with no value")
                        self.forumUsers[seq] = "No"
                    else:
                        # print("else", seq)
                        self.forumUsers[seq] = tree2.xpath(self.xPath1 + str(seq) + self.user1xpath2)
                except IndexError as e:
                    print("IndexError occurred at Last Index seq in Users was", seq)
                    self.forumUsers[seq] = "No"
                    continue
                except:
                    print("Except Seq", seq)
            print("In the End, Users are", self.forumUsers)

            print("Forum Dates")
            # forumDates = {}
            for seq in range(startIndexNextPage, len(self.Date1)+startIndexNextPage-1-1):  # len(Date1))
                # print("seq", seq)
                try:
                    if not tree2.xpath(self.xPath1 + str(seq) + self.date1xPath2):
                        # print("if", seq)
                        # print("Invalid xpath with no value")
                        # print(self.tree.xpath(self.xPath1 + str(seq) + self.date1xPath2))
                        self.forumDates[seq] = "No"
                    else:
                        # print("else", seq)
                        self.forumDates[seq] = tree2.xpath(self.xPath1 + str(seq) + self.date1xPath2)
                except IndexError as e:
                    print("IndexError occurred at Last Index seq in Dates was", seq)
                    self.forumDates[seq] = "No"
                    continue
                except:
                    print("Dates:Except Seq", seq)
            # print("In the End, Dates are", Dates)

            print("Forum Locations")
            # forumLocation = {}
            for seq in range(startIndexNextPage, len(Date2)+startIndexNextPage-1-1):
                try:
                    if not tree2.xpath(self.xPath1 + str(seq) + self.location1xPath2):  # == '':
                        self.forumLocation[seq] = "No"
                    else:
                        self.forumLocation[seq] = tree2.xpath(self.xPath1 + str(seq) + self.location1xPath2)
                except IndexError as e:
                    print("IndexError occurred at Last Index seq in location1 was", seq)
                    self.forumLocation[seq] = "No"
                    continue
                except:
                    print("UnExepected in location", seq)

            print("Forum posts1")
            # forumPosts = {}
            for seq in range(startIndexNextPage, len(Date2)+startIndexNextPage-1-1):
                try:
                    if not tree2.xpath(self.xPath1 + str(seq) + self.posts1xPath2):  # == '':
                        # print("if", seq)
                        # print("Invalid xpath with no value")
                        self.forumPosts[seq] = "No"
                    else:
                        # print("else", seq)
                        self.forumPosts[seq] = re.sub(r'[a-zA-Z]|[A-Z]* ', '',
                                                      str(tree2.xpath(self.xPath1 + str(seq) + self.posts1xPath2)))
                except IndexError as e:
                    print("Last Index seq was", seq)
                    self.forumPosts[seq] = "No"
                    continue
                except:
                        print("Except Seq", seq)

            print("Forum reviews2")
            # forumReviews = {}
            for seq in range(startIndexNextPage, len(Date2)+startIndexNextPage-1-1):  # len(Date1))
                # print("seq", seq)
                try:
                    if not tree2.xpath(self.xPath1 + str(seq) + self.reviewsxPath2):  # == '':
                        # print("if", seq)
                        # print("Invalid xpath with no value")
                        self.forumReviews[seq] = "No"
                    else:
                        # print("else", seq)
                        self.forumReviews[seq] = re.sub(r'[a-zA-Z]|[A-Z]* ', '',
                                                        str(tree2.xpath(self.xPath1 + str(seq) + self.reviewsxPath2)))
                except IndexError as e:
                    print("Last Index seq was", seq)
                    self.forumReviews[seq] = "No"
                    continue
                except:
                    print("Except Seq", seq)
            print("Forum replies2")
            # forumReplies = {}
            for seq in range(startIndexNextPage, len(Date2)+startIndexNextPage-1-1):
                try:
                    if not self.Rep2[seq-startIndexNextPage]:
                        print("if", seq)
                        print("Invalid xpath with no value")
                        self.forumReplies[seq] = "No"
                    else:
                        # print("else", seq)
                        self.forumReplies[seq] = self.Rep2[seq-startIndexNextPage]
                except IndexError as e:
                    print("IndexError occurred at Last Index seq in replies1 was", seq)
                    self.forumReplies[seq] = "No"
                    continue
                except:
                    print("Except Seq", seq)
            print("Insertion in excel started")
            for index, colnumber in zip((range(startIndexNextPage, len(Date2)+startIndexNextPage-1-1)), range(71, 71 + (len(Date2) * 6), 6)):
                try:
                    print("formuReplyPost", index, colnumber, self.forumUsers[index], self.forumDates[index],
                          self.forumLocation[index], self.forumPosts[index], self.forumReviews[index],
                          self.forumReplies[index], sep="|", end="\n")
                    worksheet.cell(row=self.seqIndex, column=colnumber + 0).value = (
                    re.sub(r"[\[\'*\'\]]", '', str(self.forumUsers[index])))
                    worksheet.cell(row=self.seqIndex, column=colnumber + 1).value = (
                    re.sub(r"[\[\'*\'\]]", '', str(self.forumDates[index])))
                    worksheet.cell(row=self.seqIndex, column=colnumber + 2).value = (
                    re.sub(r"[\[\'*\'\]]", '', str(self.forumLocation[index])))
                    worksheet.cell(row=self.seqIndex, column=colnumber + 3).value = (
                    re.sub(r"[\[\'*,*\'\]]", '', str(self.forumPosts[index])))
                    worksheet.cell(row=self.seqIndex, column=colnumber + 4).value = (
                    re.sub(r"[\[\'*\'\]]", '', str(self.forumReviews[index])))
                    worksheet.cell(row=self.seqIndex, column=colnumber + 5).value = (
                    re.sub(r'[\[\]]', '', str(self.forumReplies[index])))
                    # print("formuReplyPost", index, colnumber, self.forumUsers[index], self.forumDates[index],
                    #       self.forumLocation[index], self.forumPosts[index], self.forumReviews[index],
                    #       self.forumReplies[index], sep="|", end="\n")
                except IndexError:
                    print("IndexErrorException unhandled at index %d while insertion" % (index))
                    continue
                except KeyError as KE:
                    print("Dictionay cannot reference this key at index:%d", index, KE)


#for rownumber in range(2, len(website)):
for rownumber in range(2, len(website)):
    print("Rownumber value=%d" % (rownumber))
    print("Website", website[rownumber - 1])
    TripAdvisorXpathObj = TripAdvisorXpathForUser1(rownumber)
    TripAdvisorXpathObj.firstPost()
    time.sleep(2)
    TripAdvisorXpathObj.forumReplyPost()
    time.sleep(2)
    TripAdvisorXpathObj.NextPageforumReplyPost()

workbook.save(excelPath)
